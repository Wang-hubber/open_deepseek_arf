"""kb_ingest -- extract text, summarize, and index a document into the knowledge base."""
import json
import os
import re
import hashlib
from pathlib import Path
from datetime import datetime

import pdfplumber
import openpyxl
import requests


DEEPSEEK_URL = "https://api.deepseek.com/v1/chat/completions"
DEEPSEEK_KEY = os.environ.get("DEEPSEEK_API_KEY", "REDACTED_API_KEY")

INDEX_PATH = "kb/index.json"
TEXTS_DIR = "kb/texts"


def _call_deepseek(prompt: str, max_tokens: int = 1024, temperature: float = 0.2) -> str:
    """Call DeepSeek API and return the response text."""
    headers = {
        "Authorization": f"Bearer {DEEPSEEK_KEY}",
        "Content-Type": "application/json",
    }
    payload = {
        "model": "deepseek-chat",
        "messages": [
            {"role": "system", "content": "你是一个专业的汽车售后技术文档处理助手。请简洁、准确地输出结果。"},
            {"role": "user", "content": prompt},
        ],
        "max_tokens": max_tokens,
        "temperature": temperature,
    }
    resp = requests.post(DEEPSEEK_URL, json=payload, headers=headers, timeout=120)
    resp.raise_for_status()
    return resp.json()["choices"][0]["message"]["content"]


def _extract_pdf(file_path: str) -> str:
    """Extract full text from a PDF file."""
    parts = []
    with pdfplumber.open(file_path) as pdf:
        for i, page in enumerate(pdf.pages, 1):
            text = page.extract_text()
            if text:
                parts.append(f"--- 第{i}页 ---\n{text}")
    return "\n\n".join(parts)


def _extract_excel(file_path: str) -> str:
    """Extract full text from an Excel file."""
    wb = openpyxl.load_workbook(file_path, data_only=True)
    parts = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        parts.append(f"=== 工作表: {sheet_name} ===")
        rows_text = []
        for row in ws.iter_rows(values_only=True):
            row_str = " | ".join([str(c) if c is not None else "" for c in row])
            if row_str.strip():
                rows_text.append(row_str)
        parts.append("\n".join(rows_text))
    wb.close()
    return "\n\n".join(parts)


def _classify_document(text: str, filename: str) -> str:
    """Classify document type based on content and filename."""
    prompt = f"""根据文件名和内容片段，判断该文档属于以下哪一类（只输出类别名称）：
- 维修手册
- TPI（产品技术信息）
- 维修案例
- 保修政策
- 其他

文件名：{filename}

内容片段（前1500字）：
{text[:1500]}

类别："""
    result = _call_deepseek(prompt, max_tokens=50).strip()
    return result


def _generate_summary(text: str, filename: str) -> str:
    """Generate a concise summary of the document."""
    # Truncate if too long (DeepSeek context window is large enough)
    content = text[:8000] if len(text) > 8000 else text
    prompt = f"""为以下汽车售后技术文档生成摘要（200-400字），包含：
1. 文档主题
2. 涉及车型/部件
3. 关键问题或信息点
4. 适用场景

文件名：{filename}

文档内容：
{content}

摘要："""
    result = _call_deepseek(prompt, max_tokens=600).strip()
    return result


async def execute(file_path: str) -> dict:
    try:
        file_full_path = Path(file_path)

        if not file_full_path.exists():
            return {"error": f"文件不存在: {file_path}"}

        ext = file_full_path.suffix.lower()
        if ext == ".pdf":
            text = _extract_pdf(str(file_full_path))
        elif ext in (".xlsx", ".xls"):
            text = _extract_excel(str(file_full_path))
        else:
            return {"error": f"不支持的文件格式: {ext}，仅支持 PDF 和 Excel"}

        if not text or not text.strip():
            return {"error": "未能从文件中提取到文本内容"}

        # Generate doc_id from filename hash
        filename = file_full_path.name
        doc_id = "doc_" + hashlib.md5(filename.encode()).hexdigest()[:8]

        # Save extracted text
        os.makedirs(TEXTS_DIR, exist_ok=True)
        text_path = f"{TEXTS_DIR}/{doc_id}.txt"
        with open(text_path, "w", encoding="utf-8") as f:
            f.write(text)

        # Classify
        category = _classify_document(text, filename)

        # Generate summary
        summary = _generate_summary(text, filename)

        # Load or create index
        os.makedirs(os.path.dirname(INDEX_PATH), exist_ok=True)
        index = []
        if os.path.exists(INDEX_PATH):
            try:
                index = json.load(open(INDEX_PATH, "r", encoding="utf-8"))
            except (json.JSONDecodeError, FileNotFoundError):
                index = []

        # Remove existing entry with same doc_id (re-ingest)
        index = [entry for entry in index if entry.get("id") != doc_id]

        # Add new entry
        entry = {
            "id": doc_id,
            "filename": filename,
            "file_path": file_path,
            "category": category,
            "summary": summary,
            "text_path": text_path,
            "text_length": len(text),
            "ingested_at": datetime.now().isoformat(),
        }
        index.append(entry)

        with open(INDEX_PATH, "w", encoding="utf-8") as f:
            json.dump(index, f, ensure_ascii=False, indent=2)

        return {
            "ok": True,
            "doc_id": doc_id,
            "filename": filename,
            "category": category,
            "summary": summary,
            "text_length": len(text),
            "total_docs": len(index),
        }

    except Exception as exc:
        return {"error": str(exc), "detail": type(exc).__name__}
