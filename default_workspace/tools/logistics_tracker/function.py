import traceback
import re
import os
import random
from datetime import datetime, timedelta


def _log(level: str, message: str, **extra) -> None:
    """Internal logger — writes structured log entries to stderr."""
    import json, sys
    entry = {"ts": datetime.now().isoformat(), "level": level, "msg": message, **extra}
    print(json.dumps(entry, ensure_ascii=False), file=sys.stderr)


def _detect_carrier(tracking_number: str) -> str:
    """根据运单号自动识别快递公司"""
    tn = tracking_number.upper().strip()

    # 顺丰 — SF开头+12位数字，或纯15位数字
    if tn.startswith("SF") or (tn.isdigit() and len(tn) == 12):
        return "顺丰"

    # 京东 — JD开头
    if tn.startswith("JD") or tn.startswith("JDX"):
        return "京东"

    # 极兔 — JT开头
    if tn.startswith("JT") or tn.startswith("J&T"):
        return "极兔"

    # 德邦 — DP开头
    if tn.startswith("DP"):
        return "德邦"

    # 邮政EMS — 以E开头+字母+数字
    if re.match(r'^E[A-Z]\d', tn):
        return "邮政EMS"

    # 中通 — 数字以73/75/76/77/78开头，或ZT开头
    if tn.startswith("ZT") or re.match(r'^7[3-8]\d', tn):
        return "中通"

    # 圆通 — YT开头，或数字以10~14开头
    if tn.startswith("YT") or re.match(r'^1[0-4]\d', tn):
        return "圆通"

    # 申通 — ST开头，或数字以36~39开头
    if tn.startswith("ST") or re.match(r'^3[6-9]\d', tn):
        return "申通"

    # 韵达 — YD开头，或数字以43~45开头
    if tn.startswith("YD") or re.match(r'^4[3-5]\d', tn):
        return "韵达"

    # 百世 — BS开头
    if tn.startswith("BS"):
        return "百世"

    # 默认
    return "其他"


def _generate_mock_tracking(carrier: str, days_ago_range: tuple = (1, 14)) -> dict:
    """生成模拟的物流追踪信息"""
    now = datetime.now()
    start_days = random.randint(days_ago_range[0], days_ago_range[1])

    # 随机决定物流阶段
    stages = ["已揽收", "运输中", "派送中", "已签收", "疑难"]
    weights = [0.2, 0.3, 0.2, 0.25, 0.05]
    stage = random.choices(stages, weights=weights, k=1)[0]

    cities = [
        "广州", "深圳", "北京", "上海", "杭州", "武汉", "成都",
        "南京", "苏州", "重庆", "西安", "长沙", "郑州", "东莞"
    ]
    hubs = ["分拣中心", "中转站", "营业部", "集散中心"]

    traces = []
    current_time = now - timedelta(days=start_days)

    # 揽收
    traces.append({
        "time": current_time.strftime("%Y-%m-%d %H:%M:%S"),
        "status": "已揽收",
        "location": f"{random.choice(cities)} {random.choice(hubs)}"
    })

    # 运输中
    num_transit = random.randint(1, 4)
    for i in range(num_transit):
        current_time += timedelta(hours=random.randint(4, 24))
        if current_time > now:
            break
        city = random.choice(cities)
        traces.append({
            "time": current_time.strftime("%Y-%m-%d %H:%M:%S"),
            "status": "运输中",
            "location": f"已到达 {city}{random.choice(hubs)}"
        })

    # 最终状态
    if stage == "派送中":
        current_time += timedelta(hours=random.randint(1, 6))
        if current_time <= now:
            traces.append({
                "time": current_time.strftime("%Y-%m-%d %H:%M:%S"),
                "status": "派送中",
                "location": f"快递员正在派送，预计今日送达"
            })
    elif stage == "已签收":
        current_time += timedelta(hours=random.randint(2, 12))
        if current_time <= now:
            traces.append({
                "time": current_time.strftime("%Y-%m-%d %H:%M:%S"),
                "status": "已签收",
                "location": f"已签收（{random.choice(['前台', '本人', '门卫', '快递柜'])}）"
            })
    elif stage == "疑难":
        traces.append({
            "time": (now - timedelta(hours=2)).strftime("%Y-%m-%d %H:%M:%S"),
            "status": "疑难",
            "location": f"收件人电话无法接通，待二次派送"
        })

    latest = traces[-1] if traces else {"time": "", "status": "未知", "location": ""}
    return {
        "status": latest["status"],
        "latest_trace": f"{latest['time']} {latest['location']}",
        "trace_detail": traces,
        "tracking_days": start_days
    }


def _read_excel(file_path: str) -> list:
    """读取Excel文件，解析订单列表"""
    try:
        import openpyxl
    except ImportError:
        return {"error": "缺少openpyxl库，请执行: pip install openpyxl"}

    if not os.path.exists(file_path):
        return {"error": f"文件不存在: {file_path}"}

    try:
        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
        ws = wb.active

        # 读取表头
        headers = []
        for cell in next(ws.iter_rows(min_row=1, max_row=1, values_only=True)):
            headers.append(str(cell).strip() if cell else "")

        # 查找列索引
        col_map = {}
        for idx, h in enumerate(headers):
            h_lower = h.lower().replace(" ", "").replace("_", "")
            if h_lower in ("订单号", "order_id", "订单id", "订单编号"):
                col_map["order_id"] = idx
            elif h_lower in ("运单号", "tracking_number", "tracking_no", "快递单号", "物流单号"):
                col_map["tracking_number"] = idx
            elif h_lower in ("快递公司", "carrier", "快递", "物流公司"):
                col_map["carrier"] = idx

        if "order_id" not in col_map:
            return {"error": "Excel文件中未找到'订单号'列，请确保列名为：订单号 / order_id"}
        if "tracking_number" not in col_map:
            return {"error": "Excel文件中未找到'运单号'列，请确保列名为：运单号 / tracking_number"}

        orders = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            order_id = str(row[col_map["order_id"]]).strip() if row[col_map["order_id"]] else ""
            if not order_id or order_id == "None":
                continue
            tracking_number = str(row[col_map["tracking_number"]]).strip() if row[col_map["tracking_number"]] else ""
            carrier = ""
            if "carrier" in col_map and row[col_map["carrier"]]:
                carrier = str(row[col_map["carrier"]]).strip()

            orders.append({
                "order_id": order_id,
                "tracking_number": tracking_number,
                "carrier": carrier
            })

        wb.close()
        return orders

    except Exception as e:
        return {"error": f"读取Excel失败: {str(e)}"}


def _write_excel(file_path: str, orders: list, results: list) -> str:
    """将查询结果写入新的Excel文件"""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        return {"error": "缺少openpyxl库，请执行: pip install openpyxl"}

    try:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "物流查询结果"

        # 定义样式
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # 写入表头
        headers = ["订单号", "运单号", "快递公司", "物流状态", "最新轨迹", "查询时间"]
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border

        # 写入数据
        query_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        for row_idx, (order, result) in enumerate(zip(orders, results), 2):
            ws.cell(row=row_idx, column=1, value=order.get("order_id", "")).border = thin_border
            ws.cell(row=row_idx, column=2, value=order.get("tracking_number", "")).border = thin_border
            ws.cell(row=row_idx, column=3, value=order.get("carrier", "")).border = thin_border
            ws.cell(row=row_idx, column=4, value=result.get("status", "")).border = thin_border
            ws.cell(row=row_idx, column=5, value=result.get("latest_trace", "")).border = thin_border
            ws.cell(row=row_idx, column=6, value=query_time).border = thin_border

        # 调整列宽
        col_widths = [20, 25, 15, 15, 50, 22]
        for i, width in enumerate(col_widths, 1):
            ws.column_dimensions[chr(64 + i)].width = width

        # 冻结首行
        ws.freeze_panes = "A2"

        wb.save(file_path)
        return file_path

    except Exception as e:
        return {"error": f"写入Excel失败: {str(e)}"}


def execute(excel_path: str = None, orders: list = None, output_path: str = None) -> dict:
    """
    批量查询物流状态

    支持两种输入方式：
    1. excel_path - 读取Excel文件中的订单
    2. orders - 直接传入订单列表

    自动识别快递公司，模拟查询物流状态，结果可导出到Excel。
    """
    _log("INFO", "execute called",
         excel_path=excel_path,
         orders_count=len(orders) if orders else 0,
         output_path=output_path)

    try:
        # ---- 校验输入 ----
        if not excel_path and not orders:
            return {"error": "请提供excel_path（Excel文件路径）或orders（订单列表）"}

        # ---- 获取订单列表 ----
        order_list = []
        if excel_path:
            result = _read_excel(excel_path)
            if isinstance(result, dict) and "error" in result:
                return result
            order_list = result
            _log("INFO", f"从Excel读取到 {len(order_list)} 个订单", excel_path=excel_path)
        elif orders:
            order_list = orders

        if not order_list:
            return {"error": "未读取到任何订单数据"}

        # ---- 逐单查询 ----
        results = []
        for order in order_list:
            tracking_number = str(order.get("tracking_number", "")).strip()
            if not tracking_number:
                results.append({
                    "status": "无效单号",
                    "latest_trace": "运单号为空"
                })
                continue

            # 识别快递公司
            carrier = order.get("carrier", "").strip()
            if not carrier:
                carrier = _detect_carrier(tracking_number)
                order["carrier"] = carrier

            # 查询物流（模拟）
            tracking_info = _generate_mock_tracking(carrier)

            _log("INFO", "查询结果",
                 order_id=order.get("order_id", ""),
                 tracking_number=tracking_number,
                 carrier=carrier,
                 status=tracking_info["status"])

            results.append({
                "status": tracking_info["status"],
                "latest_trace": tracking_info["latest_trace"],
                "trace_detail": tracking_info["trace_detail"],
                "tracking_days": tracking_info["tracking_days"]
            })

        # ---- 输出结果 ----
        summary = []
        for order, result in zip(order_list, results):
            summary.append({
                "order_id": order.get("order_id", ""),
                "tracking_number": order.get("tracking_number", ""),
                "carrier": order.get("carrier", ""),
                "status": result["status"],
                "latest_trace": result["latest_trace"]
            })

        output = {
            "ok": True,
            "total": len(order_list),
            "results": summary
        }

        # 如果提供了output_path或excel_path，写入Excel文件
        if output_path:
            write_result = _write_excel(output_path, order_list, results)
            if isinstance(write_result, dict) and "error" in write_result:
                output["export_error"] = write_result["error"]
            else:
                output["output_file"] = write_result
                _log("INFO", "结果已导出到Excel", output_path=output_path)
        elif excel_path:
            # 默认在同目录生成结果文件
            base_dir = os.path.dirname(os.path.abspath(excel_path))
            base_name = os.path.splitext(os.path.basename(excel_path))[0]
            default_output = os.path.join(base_dir, f"{base_name}_物流结果.xlsx")
            write_result = _write_excel(default_output, order_list, results)
            if isinstance(write_result, dict) and "error" in write_result:
                output["export_error"] = write_result["error"]
            else:
                output["output_file"] = default_output
                _log("INFO", "结果已导出到Excel", output_path=default_output)

        _log("INFO", "execute succeeded", total=len(order_list))
        return output

    except Exception as exc:
        _log("ERROR", str(exc),
             traceback=traceback.format_exc().split("\n")[-3:],
             inputs={
                 "excel_path": excel_path,
                 "orders_count": len(orders) if orders else 0,
                 "output_path": output_path
             })
        return {"error": str(exc), "detail": str(type(exc).__name__)}
